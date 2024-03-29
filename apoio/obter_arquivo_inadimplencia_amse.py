import openpyxl as xlsx
from classes import documentos_inadimplente as di
import PySimpleGUI as pg

def obter(arquivo, ignorar=True):
    planilha = xlsx.load_workbook(arquivo)
    
    aba = planilha.active

    ultima_linha = aba.max_row
    ultima_coluna = aba.max_column

    dados_linha = []

    dados_documentos_inadimplentes = []

    erro_encontrado = False

    if ignorar:
        pular_primeira_linha = 1
    else:
        pular_primeira_linha = 0

    for linha in range(1 + pular_primeira_linha, ultima_linha + 1):
        dados_linha.append(linha)
       
        for coluna in range(1, ultima_coluna + 1):
            if aba.cell(row=linha, column=coluna).value == None or \
               aba.cell(row=linha, column=1).value == ' ' or \
               aba.cell(row=linha, column=1).value == '':
                pg.popup('Arquivo excel com dados faltando.\n\nFavor verificar linha ' + str(linha))
                erro_encontrado = True
                break
            else:
                dados_linha.append(aba.cell(row=linha, column=coluna).value)

        dados_documentos_inadimplentes.append(dados_linha)
        dados_linha = []

    if erro_encontrado:
        return None

    if len(dados_documentos_inadimplentes) != 0:
        dados_documentos_inadimplentes.sort()
        documentos_inadimplentes = di.documento.criar(dados_documentos_inadimplentes)
        return documentos_inadimplentes
    else:
        pg.popup('Arquivo excel sem dados. Favor verificar')
        return None

    
