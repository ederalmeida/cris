import openpyxl as xlsx

def importar(ignorar=True):
    
    linha_planilha = []

    dados_planilha = []
    
    arquivo_inadimplentes = xlsx.load_workbook('inadimplentes.xlsx')
    
    aba_inadimplentes = arquivo_inadimplentes.get_sheet_by_name('Planilha1')
    
    ultima_linha = aba_inadimplentes.max_row
    
    ultima_coluna = aba_inadimplentes.max_column

    if ignorar:
        primeira_linha = 2
    else:
        primeira_linha = 1

    for linha in range(primeira_linha, ultima_linha +1):
        for coluna in range(1, ultima_coluna + 1):
            linha_planilha.append(aba_inadimplentes.cell(row=linha, column=coluna).value)
        dados_planilha.append(linha_planilha)
        linha_planilha = []
    
    return dados_planilha